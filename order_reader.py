import io
import re
from pdfminer.converter import TextConverter
from pdfminer.pdfinterp import PDFPageInterpreter
from pdfminer.pdfinterp import PDFResourceManager
from pdfminer.pdfpage import PDFPage
import openpyxl
import os.path
from dotenv import load_dotenv
import os

load_dotenv()
PRICE_LIST = os.getenv('PRICE_LIST')
FILENAME = os.getenv('FILENAME')
KEYS = ['Артикул',
        'Наименование',
        'Полная стоимость за штуку',
        'Стоимость со скидкой за штуку',
        'Количество',
        'Стоимость со скидкой',
        'Скидка',
        'Заказ поставщика',
        'Дата размещения заказа',
        'Номер подтверждения заказа',
        'Дата подтверждения заказа',
        'Готовность',
        ]


def extract_text_from_pdf(pdf_path):
    resource_manager = PDFResourceManager()
    fake_file_handle = io.StringIO()
    converter = TextConverter(resource_manager, fake_file_handle)
    page_interpreter = PDFPageInterpreter(resource_manager, converter)

    with open(pdf_path, 'rb') as fh:
        for page in PDFPage.get_pages(fh,
                                      caching=True,
                                      check_extractable=True):
            page_interpreter.process_page(page)

        text = fake_file_handle.getvalue()

    # close open handles
    converter.close()
    fake_file_handle.close()

    if text:
        return text


def get_names_and_full_prices_for_order_articles(sheet, article):
    for row in sheet.iter_rows(1):
        for cell in row:
            if str(cell.value) == str(article):
                name = sheet.cell(row=cell.row, column=3).value
                full_price = sheet.cell(row=cell.row, column=5).value
                return name, full_price
    return 'Нет такой позиции в прайсе', 0


def get_orders_from_text(text, sheet):
    # Артикулы вида
    # E123131
    # EK123121321
    # L133213
    # 12313V121
    # 12313S1323
    # 1213A012
    # 121313T13213
    # 12311TE121
    order_articles = re.findall(r'000\d\d(E?K?L?\d+[vVsSaAtT]?[eE]?\d*) *(\d+)', text)
    orders = []
    index = 0
    for article, quantity in order_articles:
        quantity = int(quantity)
        price_string = re.findall(r'(\d*.?\d*.?\d\d) [CV][oa][uc][nu][tu][rm][ym]', text)[index]
        price = price_string.split('.')
        price = ''.join(price)
        price = price.split(',')
        total_price = float('.'.join(price))
        supplier_order = re.findall(r'Your purchase order: (\d*)', text)[0]
        supplier_order_date = re.search(r'(\d+.\d+.\d+)', text)[0]
        order_confirmation = re.findall(r'_(\d*)127793(\d+.\d+.\d+)', text)[0][0]
        order_confirmation_date = re.findall(r'_(\d*)127793(\d+.\d+.\d+)', text)[0][1]
        ready_week = re.findall(r'(\d\d/[12][90])', text)[index]
        name, full_price = get_names_and_full_prices_for_order_articles(sheet, article)
        discount = 0
        if full_price:
            discount = (full_price * quantity - total_price) / (full_price * quantity) * 100
        order = {
            KEYS[0]: article,
            KEYS[1]: name,
            KEYS[2]: full_price,
            KEYS[3]: total_price / quantity,
            KEYS[4]: quantity,
            KEYS[5]: total_price,
            KEYS[6]: discount,
            KEYS[7]: supplier_order,
            KEYS[8]: supplier_order_date,
            KEYS[9]: order_confirmation,
            KEYS[10]: order_confirmation_date,
            KEYS[11]: ready_week,
        }
        orders.append(order)
        index += 1
    return orders


# Create the workbook and sheet for Excel
def create_xls(filename):
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    column = 1
    for key in KEYS:
        sheet.cell(row=1, column=column, value=key)
        column += 1
    workbook.save(filename=filename)


def add_order_to_xls(filename, orders):
    workbook = openpyxl.load_workbook(filename)
    sheet = workbook.active
    row = sheet.max_row + 1
    for order in orders:
        column = 1
        for key, value in order.items():
            sheet.cell(row=row, column=column, value=value)
            column += 1
        row += 1
    workbook.save(filename=filename)


def main(filename):
    pdf_paths = os.listdir()
    if not os.path.isfile(filename):
        create_xls(filename)
    workbook = openpyxl.load_workbook(PRICE_LIST, read_only=True)
    sheet = workbook.active
    for pdf_path in pdf_paths:
        if '.pdf' in pdf_path:
            print(pdf_path)
            text = extract_text_from_pdf(pdf_path)
            print(text)
            order = get_orders_from_text(text, sheet)
            add_order_to_xls(filename, order)
            os.remove(pdf_path)


if __name__ == '__main__':
    main(FILENAME)
